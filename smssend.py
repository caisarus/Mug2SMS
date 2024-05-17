import openpyxl
from datetime import datetime, timedelta

def search_for_yesterday_date(file_name, sheet_name):
    # Load the workbook and select the sheet
    print(f"Loading workbook '{file_name}'...")
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    print(f"Sheet '{sheet_name}' selected.")

    # Get yesterday's date in the desired format (e.g., '15.05.2024')
    yesterday_date_str = (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")
    print(f"Formatted yesterday's date: {yesterday_date_str}")

    # Iterate through all the cells in the sheet
    print("Starting search for yesterday's date in the sheet...")
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            cell_value = cell.value
            if isinstance(cell_value, datetime):
                cell_value = cell_value.strftime("%d.%m.%Y")
            print(f"Checking cell {cell.coordinate} (Row: {row_idx}, Column: {col_idx}) with value: {cell_value}")
            # Check if the cell value is a string and matches the formatted yesterday's date
            if isinstance(cell_value, str) and cell_value == yesterday_date_str:
                print(f"Yesterday's date found in cell {cell.coordinate} with value: {cell_value}")
                # Get the next cell in the same row
                next_column_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                print(f"Value in the next column: {next_column_value}")
                
                # Search for the next_column_value in the first 17 rows, in the second column
                print(f"Searching for '{next_column_value}' in the first 17 rows of the second column...")
                for search_row_idx in range(1, 18):
                    search_cell_value = sheet.cell(row=search_row_idx, column=2).value
                    print(f"Checking cell at Row: {search_row_idx}, Column: 2 with value: {search_cell_value}")
                    if search_cell_value == next_column_value:
                        result_value = sheet.cell(row=search_row_idx, column=3).value
                        if isinstance(result_value, (int, float)):
                            result_value = str(result_value)
                        if isinstance(result_value, str) and result_value.isdigit():
                            result_value = f"+40{result_value}"
                        print(f"Value found in row {search_row_idx}, column 3 (phone number with +40): {result_value}")
                        return
                
                print(f"Value '{next_column_value}' not found in the first 17 rows of the second column.")
                return

    print("Yesterday's date not found in the sheet.")

# File and sheet details
file_name = "Mug2.xlsx"
sheet_name = "Ordine scos gunoi"

# Call the function
search_for_yesterday_date(file_name, sheet_name)

